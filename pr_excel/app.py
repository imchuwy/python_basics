'''
3h58m
pypi.org Directory PyPI has many many more packages which people have published.
In terminal. Write 'pip install [package]' here we use openpyxl
You can find these installed packages in prokect panel > external library > python3 > site-packages
'''

import openpyxl as xl
from openpyxl.chart import BarChart, Reference #From package openpyxl and module chart. We import class BarChart and Reference


wb = xl.load_workbook(r'C:\Users\oscar\Documents\MyFiles\Code\python\pr_excel\transactions.xlsx')
sheet = wb['Sheet1'] #Dictate sheet
cell = sheet['a1'] #Dictate cell
#print(cell.value)

for row in range(2, sheet.max_row + 1): #To dynamically find how many rows there are
    cell = sheet.cell(row,3)
    adj_price = cell.value * 0.9 #Use .value method to show value. Otherwise we read the object
    adj_price_cell = sheet.cell(row, 4) #Setting variable to row 4
    adj_price_cell.value = adj_price #Adding the new value to row 4


values = Reference(sheet
          ,min_row = 2
          ,max_row = sheet.max_row #From row 2 to max
          ,min_col = 4
          ,max_col = 4  #For column 4
          )

chart = BarChart() #Calling method
chart.add_data(values) #Adding values defined above
sheet.add_chart(chart, 'e2') #Input chart in e2

#Save in new location
wb.save(r'C:\Users\oscar\Documents\MyFiles\Code\python\pr_excel\transactions2.xlsx')


